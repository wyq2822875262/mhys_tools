from openpyxl import load_workbook
from bs4 import BeautifulSoup
import time, lxml, re, json, os
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models
from fuzzywuzzy import fuzz

# =======================================参数设置区====================================
data = {}  # 题库字典
base64_list = []  # 图片base64
txt_list = []  # 题目文字数组
last_list = []  # 答案数组
apiId = "腾讯id"
apiPwd = "腾讯Token"
# ====================================================================================
# 导入题库 最终以字典的形式存储
wb = load_workbook('answer.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
for row in sheet.iter_rows(min_row=2, max_col=2):
    result1 = row[0].value
    result2 = row[1].value
    data[result1] = result2
print("题库导入成功!")
# 等待网页保存,保存后结束循环
IsExist = False;
while True:
    if os.path.isfile(r'D:\index.html'):
        break
    else:
        print("未检测到index.html文件,休眠5s")
        time.sleep(5)
# 使用bs4解析html 获取所有图片的 imagebase64
# 解析html文件，将img标签提取出来
print("开始解析HTML")
html = open(r'D:\index.html', mode='r', encoding='utf8')
soup = BeautifulSoup(html, 'lxml')
img_list = soup.find_all('img')  # 题目图片数组
# 由于存在51张图片，第一张为警告图片，故将第一张删除
print("开始提取图片base64")
del img_list[0]
# 使用正则提取出所有的base64,并存入题目文字数组
# 创建正则表达式规则
pat = re.compile(r'<img src="(.*?)" style="margin-top:4px"/>')
for i in img_list:
    num = img_list.index(i)
    base64_list.append((re.findall(pat, str(img_list[num])))[0])
print("提取成功")
os.remove(r'D:\index.html')  # 删除文件


# 方法 调用腾讯api 将图片转换成文字,返回值为string
def Getimgtxt(img64):
    try:
        cred = credential.Credential(apiId, apiPwd)
        httpProfile = HttpProfile()
        httpProfile.endpoint = "ocr.tencentcloudapi.com"

        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = ocr_client.OcrClient(cred, "ap-shanghai", clientProfile)

        req = models.GeneralAccurateOCRRequest()
        params = {
            "ImageBase64": img64
        }
        req.from_json_string(json.dumps(params))

        resp = client.GeneralAccurateOCR(req)
        data = json.loads(resp.to_json_string())
        return data["TextDetections"][0]["DetectedText"]
    except TencentCloudSDKException as err:
        print(err)


# 循环 调用api将图片转换成文字，并存入列表
print("开始图片转文字")
for i in base64_list:
    txt_list.append(Getimgtxt(i))
    print("第{}题转换成功".format(base64_list.index(i) + 1))
temp_list = []  # 临时数组，用于存放为找到的答案
print("开始查找答案")
for i in txt_list:  # 查找答案
    num = txt_list.index(i)
    num += 1
    for key in data.keys():
        result_last = fuzz.ratio(i, key)
        if result_last > 80:
            print("成功匹配第{}题".format(num))
            last_list.append("第{}题: {}".format(str(num), data[key]))
            continue
# 存储答案
print("存储数据!")
f = open(r'D:\题目.txt')
for i in txt_list:
    f.write(i)
f.close()

f = open(r'D:\答案.txt')
for i in last_list:
    f.write(i)
f.close()
# 判断文件是否保存成功
file_path = ['题目', '答案']
for i in file_path:
    if os.path.isfile(r"D:\{}.txt".format(i)):
        print("{}保存成功".format(i))
    else:
        print("{}保存失败".format(i))
for i in last_list:  # 输出答案
    print(i)
print("共匹配到{}题,答案存储在D:答案.txt中,题目存储在D:题目.txt中!".format(len(last_list)))
# 未实现的功能: 钉钉机器人发送消息,输出具体错题
